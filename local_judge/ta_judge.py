#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2020 Huang Po-Hsuan

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

from .version import __version__

import sys

if sys.version_info < (3,):
    raise ImportError(
        "You are running local-judge {} on Python 2\n".format(__version__)
        + "Please use Python 3"
    )

import argparse
import configparser
import os
from openpyxl import load_workbook
from zipfile import ZipFile
import rarfile
import tarfile
from glob import glob as globbing
import re
from collections import namedtuple
import logging
import multiprocessing
import signal
import time
from . import judge

Student = namedtuple("Student", ("id", "zip_type", "zip_path", "extract_path"))


class TaJudge:
    def __init__(self, ta_config):
        try:
            self._config = ta_config
            self.students_zip_container = self._config["StudentsZipContainer"]
            self.students_pattern = self._config["StudentsPattern"]
            self.update_student_pattern = self._config["UpdateStudentPattern"]
            self.students_extract_dir = self._config["StudentsExtractDir"]
            self.extract_afresh = self._config["ExtractAfresh"]
            self.students_zips = globbing(self.students_zip_container + os.sep + "*")

            # Parse the students' id and sort them
            self.students = self._parse_students()
            self.students.sort(key=lambda x: x.id)
        except KeyError as e:
            print("[ERROR] " + str(e) + " field was not found in config file.")
            print("Please check `judge.conf` first.")
            exit(1)
        try:
            # Create the temporary directory for output
            # Suppress the error when the directory already exists
            os.makedirs(self.students_extract_dir)
        except OSError as e:
            if e.errno != os.errno.EEXIST:
                print("[ERROR] " + str(e))
                exit(1)

    def _parse_students(self):
        """Used to parse student by the zip filename.

        The filename will be split into two parts: student id and the type of zip file (zip or rar).
        """

        regex = re.compile(self.students_pattern)
        students = []
        for student_zip_path in self.students_zips:
            try:
                match = regex.search(student_zip_path)
                # group(1): whole filename for extracted path
                # group(2): student id
                # group(3): file type (zip or rar)
                students.append(
                    Student(
                        match.group(2),
                        match.group(3),
                        os.path.abspath(student_zip_path),
                        os.path.abspath(
                            self.students_extract_dir + os.sep + match.group(1)
                        ),
                    )
                )
            except Exception as e:
                logging.error(
                    "Failed in parse stage. `"
                    + str(student_zip_path)
                    + "` did not match the file rule."
                )
        return students

    def extract_student(self, student):
        """Used to extract students' zip file."""
        if self.extract_afresh == "false":
            return
        try:
            if student.zip_type == "zip":
                z = ZipFile(student.zip_path, "r")
            elif student.zip_type == "tar":
                z = tarfile.open(student.zip_path, "r")
            elif student.zip_type == "rar":
                z = rarfile.RarFile(student.zip_path)
            else:
                logging.error(
                    str(student.id)
                    + " failed in extract stage with unknown zip type: `"
                    + student.zip_type
                    + "`"
                )
                return
            z.extractall(self.students_extract_dir)
            z.close()
        except Exception as e:
            logging.error(
                "`" + str(student.id) + "` failed in extract stage. " + str(e)
            )


def find_student_by_id(student_id, students):
    """Find the student from the students by id."""
    return [s for s in students if s.id == student_id]


def append_log_msg(ori_result, log_msg):
    if log_msg != "":
        return ori_result + [1, log_msg]
    else:
        return ori_result + [0]


def judge_one_student(
    student, all_student_results, tj: TaJudge, lj: judge.LocalJudge, skip_report=False
):
    """Judge one student and return the correctness result."""
    lj.error_handler.init_student(student.id)
    print(student.id)
    student_path = student.extract_path + os.sep
    correctness = [0] * len(lj.tests)
    correct_cnt = 0
    report_table = []
    tj.extract_student(student)
    if not os.path.isdir(student_path):
        lj.error_handler.handle(
            f"File architecture error: {str(student.id)}: {str(student_path)} not found",
            student_id=student.id,
        )
    else:
        lj.build(student_id=student.id, cwd=student_path)
        for i in range(len(lj.tests)):
            returncode, output = lj.run(
                lj.tests[i].input_filepath, student_id=student.id, cwd=student_path
            )
            accept, diff = lj.compare(
                output,
                lj.tests[i].answer_filepath,
                returncode,
                student_id=student.id,
                cwd=student_path,
            )
            if not skip_report:
                report_table.append(
                    {"test": lj.tests[i].test_name, "accept": accept, "diff": diff}
                )
            if accept:
                correct_cnt += 1
                correctness[i] = 1
            else:
                correctness[i] = 0
    # Not calcute the total score because there may be some hidden test cases
    # Use the formula of google sheet or excel to get the total score.
    # correctness.append(int(lj.score_dict[str(correct_cnt)]))
    result = append_log_msg(correctness, lj.error_handler.get_error(student.id))
    if not all_student_results is None:
        all_student_results[student.id] = result
    return {
        "student_id": student.id,
        "result": result,
        "report_table": report_table,
    }


def write_to_sheet(score_output_path, student_list_path, all_student_results, tests):
    # Judge all students
    # Init the table with the title
    book = load_workbook(student_list_path)
    sheet = book.active
    new_title = (
        ["name", "student_id"] + [t.test_name for t in tests] + ["in_log", "log_msg"]
    )
    for idx, val in enumerate(new_title):
        sheet.cell(row=1, column=idx + 1).value = val

    # Save result to excel
    for row in list(sheet.rows)[1:]:
        # Skip title row so use [1:]

        # row[0] are students' name, row[1] are IDs
        this_student_id = row[1].value

        this_student_result = [""] * len(tests)
        if not this_student_id in all_student_results.keys():
            # this_student_result.append(0)  # not submit: total score is 0
            this_student_result = append_log_msg(this_student_result, "not submit")
        else:
            this_student_result = all_student_results[this_student_id]
        for idx, test_result in enumerate(this_student_result):
            sheet.cell(row=row[1].row, column=idx + 3).value = str(test_result)

    book.save(score_output_path)


def get_args():
    """Init argparser and return the args from cli."""
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-c",
        "--config",
        help="the config file",
        type=str,
        default=os.getcwd() + os.sep + "judge.conf",
    )
    parser.add_argument(
        "-t",
        "--ta-config",
        help="the config file",
        type=str,
        default=os.getcwd() + os.sep + "ta_judge.conf",
    )
    parser.add_argument(
        "-s", "--student", help="judge only one student", type=str, default=None
    )
    parser.add_argument(
        "-v", "--verbose", help="the verbose level for report", type=int, default=0
    )
    parser.add_argument(
        "-j",
        "--jobs",
        help="number of jobs for multiprocessing",
        type=int,
        default=multiprocessing.cpu_count(),
    )
    parser.add_argument(
        "-u",
        "--update",
        help="update specific student's score by rejudgement",
        type=str,
        default=None,
    )
    return parser.parse_args()


def setup():
    signal.signal(signal.SIGINT, signal.SIG_IGN)


def main():
    print(f"local-judge: v{__version__}")
    args = get_args()
    if not os.path.isfile(args.ta_config):
        print("Config file `" + args.ta_config + "` not found.")
        exit(1)
    ta_config = configparser.ConfigParser()
    ta_config.read(args.ta_config)

    # Check if the config file is empty or not exist.
    if ta_config.sections() == []:
        print("[ERROR] Failed in config stage.")
        raise FileNotFoundError(args.ta_config)

    # logging configuration
    logging_config = {
        "filename": "ta_judge.log",
        "filemode": "a",
        "format": "%(asctime)-15s [%(levelname)s] %(message)s",
    }

    eh = judge.ErrorHandler(ta_config["Config"]["ExitOrLog"], **logging_config)
    tj = TaJudge(ta_config["TaConfig"])
    lj = judge.LocalJudge(ta_config["Config"], eh)

    if not args.student is None:
        # Assign specific student for this judgement and report to screen
        this_student_id = args.student
        tj.extract_afresh = "false"
        extract_path = re.sub(
            r"{student_id}", this_student_id, tj.update_student_pattern
        )
        student = Student(
            this_student_id,
            "none",
            "none",
            os.path.abspath(tj.students_extract_dir + os.sep + extract_path),
        )
        res_dict = judge_one_student(student, None, tj, lj, False)

        report = judge.Report(report_verbose=args.verbose, score_dict=lj.score_dict)
        report.table = res_dict["report_table"]
        report.print_report()

    elif not args.update is None:
        # Update one student's judge result
        this_student_id = args.update
        tj.extract_afresh = "false"
        extract_path = re.sub(
            r"{student_id}", this_student_id, tj.update_student_pattern
        )
        student = Student(
            this_student_id,
            "none",
            "none",
            os.path.abspath(tj.students_extract_dir + os.sep + extract_path),
        )
        res_dict = judge_one_student(student, None, tj, lj, False)
        result = res_dict["result"]
        # Load existing table
        book = load_workbook(ta_config["TaConfig"]["ScoreOutput"])
        sheet = book.active
        for row in sheet.rows:
            if row[1].value == this_student_id:
                for cell in row:
                    if sheet.cell(row=1, column=cell.column).value in [
                        "name",
                        "student_id",
                        "in_log",
                        "log_msg",
                    ]:
                        continue
                    if cell.column - 3 >= len(result):
                        break
                    cell.value = result[cell.column - 3]
                break
        book.save(ta_config["TaConfig"]["ScoreOutput"])

    elif args.jobs > 1:
        # Test phase
        with multiprocessing.Manager() as manager:
            all_student_results = manager.dict()
            pool = multiprocessing.Pool(args.jobs, setup)
            async_result_list = [
                {
                    "student_id": s.id,
                    "async_result": pool.apply_async(
                        judge_one_student, (s, all_student_results, tj, lj, True)
                    ),
                }
                for s in tj.students
            ]
            for async_result_i in async_result_list:
                try:
                    async_result_i["async_result"].get(
                        float(lj.timeout) * len(lj.tests) * 10
                    )
                except multiprocessing.TimeoutError:
                    print(async_result_i["student_id"], "total TLE skip")
                    time.sleep(0.5)
                    lj.error_handler.handle(
                        "total TLE skip", student_id=async_result_i["student_id"]
                    )

                    continue
                except KeyboardInterrupt:
                    pool.terminate()
                    pool.join()
                    sys.exit(1)

            pool.close()
            pool.join()

            write_to_sheet(
                ta_config["TaConfig"]["ScoreOutput"],
                ta_config["TaConfig"]["StudentList"],
                dict(all_student_results),
                lj.tests,
            )
    else:
        # Test in one thread
        all_student_results = {}
        empty_result = [""] * len(lj.tests)
        for student in tj.students:
            try:
                result_pack = judge_one_student(
                    student, all_student_results, tj, lj, True
                )
                all_student_results[student.id] = result_pack["result"]
            except KeyboardInterrupt:
                if input("\nReally quit? (y/n)> ").lower().startswith("y"):
                    # Ref: https://stackoverflow.com/a/18115530
                    print("Write current score to sheet")
                    break
                print(f"Skip one student: {student.id}")
                lj.error_handler.handle("skip", student_id=student.id)
                result = append_log_msg(
                    empty_result, lj.error_handler.get_error(student.id)
                )
                all_student_results[student.id] = result
                continue
        write_to_sheet(
            ta_config["TaConfig"]["ScoreOutput"],
            ta_config["TaConfig"]["StudentList"],
            all_student_results,
            lj.tests,
        )
    print("Finished")


if __name__ == "__main__":
    main()
